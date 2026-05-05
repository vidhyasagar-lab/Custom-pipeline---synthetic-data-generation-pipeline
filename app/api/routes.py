"""FastAPI API routes for the Synthetic Data Generation pipeline (LangGraph)."""

import asyncio
import io
import time
from pathlib import Path

from fastapi import APIRouter, BackgroundTasks, HTTPException, UploadFile, File
from fastapi.responses import JSONResponse, StreamingResponse

from app.models.schemas import PipelineRunRequest
from app.agents.supervisor import Supervisor
from app.core.config import get_settings
from app.core.logging_config import get_agent_logger
from app.core.progress import create_progress_tracker, get_progress_tracker, remove_progress_tracker
from app.core.experiments import ExperimentStore, Experiment, ExperimentConfig, ExperimentResult

logger = get_agent_logger("API")
router = APIRouter(prefix="/api/v1", tags=["pipeline"])

# In-memory store for pipeline runs with TTL cleanup
_pipeline_runs: dict[str, dict] = {}
_run_timestamps: dict[str, float] = {}  # run_id -> creation timestamp
RUN_TTL_SECONDS = 3600  # 1 hour TTL for completed runs
MAX_CONCURRENT_RUNS = 3

_experiment_store = ExperimentStore()

# Semaphore to limit concurrent pipeline executions
_run_semaphore = asyncio.Semaphore(MAX_CONCURRENT_RUNS)

# Singleton supervisor (avoids rebuilding graph per request)
_supervisor: Supervisor | None = None


def _get_supervisor() -> Supervisor:
    global _supervisor
    if _supervisor is None:
        _supervisor = Supervisor()
    return _supervisor


def _cleanup_stale_runs():
    """Remove completed/failed runs older than TTL."""
    now = time.time()
    stale = [
        run_id for run_id, ts in _run_timestamps.items()
        if now - ts > RUN_TTL_SECONDS
        and _pipeline_runs.get(run_id, {}).get("status") in ("completed", "failed")
    ]
    for run_id in stale:
        _pipeline_runs.pop(run_id, None)
        _run_timestamps.pop(run_id, None)
        remove_progress_tracker(run_id)
    if stale:
        logger.info(f"Cleaned up {len(stale)} stale run(s)")


ALLOWED_EXTENSIONS = {".docx", ".pdf"}


@router.post("/generate")
async def start_generation(request: PipelineRunRequest):
    """
    Start a chatbot synthetic data generation pipeline run (LangGraph).

    Executes the 5-agent pipeline:
    START → Semantic Chunker → Question Generator → Answer Generator
          → Deduplicator → Quality Validator → END
    """
    logger.set_step("generate")
    logger.info("Pipeline generation request received")
    logger.info(f"  quality_threshold={request.quality_threshold}, max_retries={request.max_retries}")

    supervisor = _get_supervisor()

    doc_paths = request.document_paths
    doc_path = request.document_path

    if not doc_paths and not doc_path:
        settings = get_settings()
        doc_files = sorted(
            list(settings.data_dir.glob("*.docx")) +
            list(settings.data_dir.glob("*.pdf"))
        )
        if not doc_files:
            raise HTTPException(
                status_code=400,
                detail="No document provided and no supported files found in data directory",
            )
        doc_paths = [str(f) for f in doc_files]
        logger.info(f"  Auto-detected {len(doc_paths)} document(s)")

    try:
        response = await supervisor.execute_pipeline(
            document_path=doc_path,
            document_paths=doc_paths,
            quality_threshold=request.quality_threshold,
            max_retries=request.max_retries,
        )
        _pipeline_runs[response["run_id"]] = response
        _run_timestamps[response["run_id"]] = time.time()
        _cleanup_stale_runs()
        logger.info(f"Pipeline completed: run_id={response['run_id']}")
        return response

    except Exception as e:
        logger.error(f"Pipeline execution failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/generate/async")
async def start_generation_async(
    request: PipelineRunRequest, background_tasks: BackgroundTasks
):
    """
    Start pipeline asynchronously. Returns run_id for status polling.
    """
    logger.set_step("generate_async")
    logger.info("Async pipeline generation request received")

    supervisor = _get_supervisor()

    doc_paths = request.document_paths
    doc_path = request.document_path

    if not doc_paths and not doc_path:
        settings = get_settings()
        doc_files = sorted(
            list(settings.data_dir.glob("*.docx")) +
            list(settings.data_dir.glob("*.pdf"))
        )
        if doc_files:
            doc_paths = [str(f) for f in doc_files]

    from uuid import uuid4
    run_id = str(uuid4())
    _pipeline_runs[run_id] = {"run_id": run_id, "status": "pending"}
    _run_timestamps[run_id] = time.time()

    # Cleanup stale runs before accepting new ones
    _cleanup_stale_runs()

    # Check active run count
    active = sum(1 for r in _pipeline_runs.values() if r.get("status") == "running")
    if active >= MAX_CONCURRENT_RUNS:
        _pipeline_runs.pop(run_id, None)
        _run_timestamps.pop(run_id, None)
        raise HTTPException(
            status_code=429,
            detail=f"Maximum concurrent runs ({MAX_CONCURRENT_RUNS}) reached. Try again later.",
        )

    progress = create_progress_tracker(run_id)

    async def _run_pipeline():
        async with _run_semaphore:
            try:
                _pipeline_runs[run_id]["status"] = "running"
                progress.status = "running"
                progress.log_message("Pipeline execution started")
                response = await supervisor.execute_pipeline(
                    document_path=doc_path,
                    document_paths=doc_paths,
                    quality_threshold=request.quality_threshold,
                    max_retries=request.max_retries,
                    progress=progress,
                )
                _pipeline_runs[run_id] = response
            except Exception as e:
                _pipeline_runs[run_id] = {
                    "run_id": run_id,
                    "status": "failed",
                    "errors": [str(e)],
                }
                progress.pipeline_failed(str(e))

    background_tasks.add_task(_run_pipeline)
    logger.info(f"Async pipeline queued: run_id={run_id}")
    return {"run_id": run_id, "status": "accepted"}


@router.get("/status/{run_id}")
async def get_pipeline_status(run_id: str):
    """Get the status and results of a pipeline run."""
    if run_id not in _pipeline_runs:
        raise HTTPException(status_code=404, detail=f"Run {run_id} not found")
    return _pipeline_runs[run_id]


@router.get("/results/{run_id}/export")
async def export_results(run_id: str, format: str = "json"):
    """Export QA triples from a completed pipeline run (JSON or Excel)."""
    if run_id not in _pipeline_runs:
        raise HTTPException(status_code=404, detail=f"Run {run_id} not found")

    run = _pipeline_runs[run_id]
    if run.get("status") != "completed":
        raise HTTPException(
            status_code=400, detail=f"Run is not completed. Status: {run.get('status')}"
        )

    if format == "json":
        return JSONResponse(
            content={
                "run_id": run["run_id"],
                "format": "conversational",
                "total_conversations": len(run.get("conversations", [])),
                "total_triples": len(run.get("qa_triples", [])),
                "quality": run.get("quality", {}),
                "summary": run.get("summary", {}),
                "retrieval_method": run.get("retrieval_method", ""),
                "conversations": run.get("conversations", []),
                "qa_triples": run.get("qa_triples", []),
            }
        )
    elif format == "excel":
        return _build_excel_response(run)
    else:
        raise HTTPException(status_code=400, detail=f"Unsupported format: {format}. Use 'json' or 'excel'.")


def _build_excel_response(run: dict) -> StreamingResponse:
    """Build an Excel workbook from pipeline results and return as streaming download."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()

    # ── Style definitions ─────────────────────────────────────────
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    good_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    warn_fill = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")
    bad_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

    def _style_header(ws, col_count):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

    def _score_fill(score):
        if score >= 0.8:
            return good_fill
        elif score >= 0.6:
            return warn_fill
        return bad_fill

    # ── Sheet 1: QA Pairs (main data) ────────────────────────────
    ws_qa = wb.active
    ws_qa.title = "QA Pairs"

    qa_headers = [
        "No.", "Question", "Answer", "Follow-up Question", "Follow-up Answer",
        "Question Type", "Difficulty", "Source File", "Quality Score",
        "Faithfulness", "Relevance", "Completeness", "Tone", "Follow-up Quality",
    ]
    ws_qa.append(qa_headers)
    _style_header(ws_qa, len(qa_headers))

    triples = run.get("qa_triples", [])
    for idx, t in enumerate(triples, 1):
        scores = t.get("validation_scores", {})
        row = [
            idx,
            t.get("question", ""),
            t.get("answer", ""),
            t.get("follow_up_q", ""),
            t.get("follow_up_a", ""),
            t.get("question_type", ""),
            t.get("difficulty", ""),
            ", ".join(t.get("retrieved_sources", [])[:2]) if t.get("retrieved_sources") else "",
            round(t.get("quality_score", 0.0), 3),
            round(scores.get("faithfulness", 0.0), 3),
            round(scores.get("relevance", 0.0), 3),
            round(scores.get("completeness", 0.0), 3),
            round(scores.get("tone", 0.0), 3),
            round(scores.get("follow_up_quality", 0.0), 3),
        ]
        ws_qa.append(row)
        row_num = idx + 1
        for col in range(1, len(qa_headers) + 1):
            ws_qa.cell(row=row_num, column=col).alignment = cell_alignment
            ws_qa.cell(row=row_num, column=col).border = thin_border
        # Color-code quality score
        ws_qa.cell(row=row_num, column=9).fill = _score_fill(t.get("quality_score", 0.0))

    # Set column widths
    qa_widths = [6, 50, 60, 50, 60, 14, 12, 25, 12, 12, 12, 14, 10, 16]
    for i, w in enumerate(qa_widths, 1):
        ws_qa.column_dimensions[ws_qa.cell(row=1, column=i).column_letter].width = w

    # Freeze header row
    ws_qa.freeze_panes = "A2"

    # ── Sheet 2: Conversations (multi-turn format) ────────────────
    ws_conv = wb.create_sheet("Conversations")
    conv_headers = ["Conv No.", "Turn", "Role", "Content", "Quality Score", "Source File", "Cross-Document"]
    ws_conv.append(conv_headers)
    _style_header(ws_conv, len(conv_headers))

    conversations = run.get("conversations", [])
    for conv_idx, conv in enumerate(conversations, 1):
        meta = conv.get("metadata", {})
        messages = conv.get("messages", [])
        for turn_idx, msg in enumerate(messages, 1):
            row = [
                conv_idx,
                turn_idx,
                msg.get("role", ""),
                msg.get("content", ""),
                round(meta.get("quality_score", 0.0), 3) if turn_idx == 1 else "",
                meta.get("source_file", "") if turn_idx == 1 else "",
                "Yes" if meta.get("cross_document") and turn_idx == 1 else "",
            ]
            ws_conv.append(row)
            row_num = ws_conv.max_row
            for col in range(1, len(conv_headers) + 1):
                ws_conv.cell(row=row_num, column=col).alignment = cell_alignment
                ws_conv.cell(row=row_num, column=col).border = thin_border

    conv_widths = [10, 8, 12, 70, 12, 25, 15]
    for i, w in enumerate(conv_widths, 1):
        ws_conv.column_dimensions[ws_conv.cell(row=1, column=i).column_letter].width = w
    ws_conv.freeze_panes = "A2"

    # ── Sheet 3: Summary ──────────────────────────────────────────
    ws_summary = wb.create_sheet("Summary")
    summary_headers = ["Metric", "Value"]
    ws_summary.append(summary_headers)
    _style_header(ws_summary, len(summary_headers))

    summary = run.get("summary", {})
    quality = run.get("quality", {})
    summary_rows = [
        ("Run ID", run.get("run_id", "")),
        ("Status", run.get("status", "")),
        ("Started At", run.get("started_at", "")),
        ("Completed At", run.get("completed_at", "")),
        ("Total Time (seconds)", run.get("total_time_seconds", "")),
        ("", ""),
        ("Documents Processed", summary.get("documents_processed", "")),
        ("Document Names", ", ".join(summary.get("document_names", []))),
        ("Total Chunks", summary.get("total_chunks", "")),
        ("Total Questions Generated", summary.get("total_questions_generated", "")),
        ("Total Answers Generated", summary.get("total_answers_generated", "")),
        ("After Deduplication", summary.get("after_dedup", "")),
        ("Final QA Pairs (Validated)", summary.get("total_triples", "")),
        ("Final Conversations", summary.get("total_conversations", "")),
        ("Cross-Document Conversations", summary.get("cross_document_conversations", "")),
        ("Rejected", summary.get("rejected", "")),
        ("", ""),
        ("Average Quality Score", quality.get("avg_quality_score", "")),
        ("Retrieval Method", run.get("retrieval_method", "")),
    ]

    criteria = quality.get("criteria_averages", {})
    if criteria:
        summary_rows.append(("", ""))
        summary_rows.append(("--- Criteria Averages ---", ""))
        for k, v in criteria.items():
            summary_rows.append((k.replace("_", " ").title(), round(v, 3)))

    for label, value in summary_rows:
        ws_summary.append([label, value])
        row_num = ws_summary.max_row
        ws_summary.cell(row=row_num, column=1).font = Font(bold=True) if label else Font()
        for col in range(1, 3):
            ws_summary.cell(row=row_num, column=col).alignment = cell_alignment
            ws_summary.cell(row=row_num, column=col).border = thin_border

    ws_summary.column_dimensions["A"].width = 35
    ws_summary.column_dimensions["B"].width = 50
    ws_summary.freeze_panes = "A2"

    # ── Sheet 4: Rejected (if any) ────────────────────────────────
    rejected = run.get("rejected_triples", [])
    if rejected:
        ws_rej = wb.create_sheet("Rejected")
        rej_headers = ["No.", "Question", "Answer", "Quality Score", "Reason"]
        ws_rej.append(rej_headers)
        _style_header(ws_rej, len(rej_headers))

        for idx, t in enumerate(rejected, 1):
            row = [
                idx,
                t.get("question", ""),
                t.get("answer", ""),
                round(t.get("quality_score", 0.0), 3),
                t.get("validation_reasoning", ""),
            ]
            ws_rej.append(row)
            row_num = idx + 1
            for col in range(1, len(rej_headers) + 1):
                ws_rej.cell(row=row_num, column=col).alignment = cell_alignment
                ws_rej.cell(row=row_num, column=col).border = thin_border

        rej_widths = [6, 50, 60, 12, 50]
        for i, w in enumerate(rej_widths, 1):
            ws_rej.column_dimensions[ws_rej.cell(row=1, column=i).column_letter].width = w
        ws_rej.freeze_panes = "A2"

    # ── Write to bytes buffer ─────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"synthetic_data_{run.get('run_id', 'export')[:8]}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/upload-document")
async def upload_document(file: UploadFile = File(...)):
    """Upload a single document to the data directory for processing."""
    logger.set_step("upload")
    settings = get_settings()

    ext = Path(file.filename).suffix.lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(status_code=400, detail=f"Unsupported file type: {ext}. Allowed: {ALLOWED_EXTENSIONS}")

    safe_filename = Path(file.filename).name
    save_path = settings.data_dir / safe_filename
    settings.data_dir.mkdir(parents=True, exist_ok=True)

    content = await file.read()

    # File size limit
    max_size = settings.max_upload_size_mb * 1024 * 1024
    if len(content) > max_size:
        raise HTTPException(status_code=413, detail=f"File too large. Max: {settings.max_upload_size_mb}MB")
    with open(save_path, "wb") as f:
        f.write(content)

    logger.info(f"Document uploaded: {safe_filename} ({len(content)} bytes)")
    return {"filename": safe_filename, "path": str(save_path), "size_bytes": len(content)}


# ═══════════════════════════════════════════════════════════════════════════════
# EXPERIMENT ENDPOINTS
# ═══════════════════════════════════════════════════════════════════════════════

@router.get("/experiments")
async def list_experiments():
    """List all experiment runs."""
    return {"experiments": _experiment_store.list_experiments()}


@router.get("/experiments/{experiment_id}")
async def get_experiment(experiment_id: str):
    """Get details of a specific experiment."""
    result = _experiment_store.get_experiment(experiment_id)
    if not result:
        raise HTTPException(status_code=404, detail=f"Experiment {experiment_id} not found")
    return result


@router.get("/experiments/compare/{id_a}/{id_b}")
async def compare_experiments(id_a: str, id_b: str):
    """Compare two experiments side by side."""
    result = _experiment_store.compare(id_a, id_b)
    if not result:
        raise HTTPException(status_code=404, detail="One or both experiments not found")
    return result


@router.delete("/experiments/{experiment_id}")
async def delete_experiment(experiment_id: str):
    """Delete an experiment record."""
    if _experiment_store.delete(experiment_id):
        return {"deleted": experiment_id}
    raise HTTPException(status_code=404, detail=f"Experiment {experiment_id} not found")


@router.post("/upload-documents")
async def upload_documents(files: list[UploadFile] = File(...)):
    """Upload multiple documents for cross-document retrieval."""
    logger.set_step("upload_multi")
    settings = get_settings()
    settings.data_dir.mkdir(parents=True, exist_ok=True)

    uploaded = []
    for file in files:
        ext = Path(file.filename).suffix.lower()
        if ext not in ALLOWED_EXTENSIONS:
            raise HTTPException(
                status_code=400,
                detail=f"Unsupported file type: {ext}. Allowed: {ALLOWED_EXTENSIONS}. Got: {file.filename}",
            )

        safe_filename = Path(file.filename).name
        save_path = settings.data_dir / safe_filename

        content = await file.read()

        # File size limit
        max_size = settings.max_upload_size_mb * 1024 * 1024
        if len(content) > max_size:
            raise HTTPException(status_code=413, detail=f"File '{safe_filename}' too large. Max: {settings.max_upload_size_mb}MB")
        with open(save_path, "wb") as f:
            f.write(content)

        uploaded.append({
            "filename": safe_filename,
            "path": str(save_path),
            "size_bytes": len(content),
        })
        logger.info(f"Document uploaded: {safe_filename} ({len(content)} bytes)")

    logger.info(f"Total documents uploaded: {len(uploaded)}")
    return {
        "uploaded": uploaded,
        "total": len(uploaded),
        "paths": [u["path"] for u in uploaded],
    }


@router.delete("/cache")
async def clear_cache():
    """Clear the Knowledge Graph cache to force rebuild on next run."""
    logger.set_step("clear_cache")
    settings = get_settings()
    cache_files = list(settings.cache_dir.glob("*.json"))

    deleted = []
    for f in cache_files:
        f.unlink()
        deleted.append(f.name)
        logger.info(f"Deleted cache file: {f.name}")

    logger.info(f"Cache cleared: {len(deleted)} files removed")
    return {"deleted": deleted, "message": f"Cleared {len(deleted)} cached files"}


@router.get("/health")
async def health_check():
    """Health check endpoint."""
    settings = get_settings()
    return {
        "status": "healthy",
        "orchestrator": "LangGraph StateGraph",
        "model": settings.azure_openai_model_name,
        "deployment": settings.azure_openai_deployment_name,
        "data_dir": str(settings.data_dir),
        "cache_dir": str(settings.cache_dir),
        "active_runs": len(_pipeline_runs),
    }


@router.get("/graph-info")
async def graph_info():
    """Get information about the LangGraph pipeline topology."""
    return {
        "framework": "LangGraph",
        "topology": "Linear StateGraph (5 agents)",
        "nodes": [
            {"name": "semantic_chunker", "agent": "Agent 1: Semantic Chunker", "tool": "Embeddings"},
            {"name": "question_generator", "agent": "Agent 2: Question Generator", "tool": "LLM"},
            {"name": "answer_generator", "agent": "Agent 3: RAG Answer Generator", "tool": "LLM + Vector Index"},
            {"name": "deduplicator", "agent": "Agent 4: Deduplicator", "tool": "Embeddings"},
            {"name": "quality_validator", "agent": "Agent 5: Quality Validator", "tool": "LLM-as-Judge"},
        ],
        "edges": [
            {"from": "START", "to": "semantic_chunker"},
            {"from": "semantic_chunker", "to": "question_generator"},
            {"from": "question_generator", "to": "answer_generator"},
            {"from": "answer_generator", "to": "deduplicator"},
            {"from": "deduplicator", "to": "quality_validator"},
            {"from": "quality_validator", "to": "END"},
        ],
    }


@router.get("/progress/{run_id}")
async def stream_progress(run_id: str):
    """Stream real-time progress events via Server-Sent Events (SSE)."""
    tracker = get_progress_tracker(run_id)
    if tracker is None:
        raise HTTPException(status_code=404, detail=f"No progress tracker for run {run_id}")

    return StreamingResponse(
        tracker.event_stream(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )


@router.get("/progress/{run_id}/logs")
async def get_progress_logs(run_id: str):
    """Get accumulated progress logs for a run."""
    tracker = get_progress_tracker(run_id)
    if tracker is None:
        raise HTTPException(status_code=404, detail=f"No progress tracker for run {run_id}")
    return {"run_id": run_id, "status": tracker.status, "logs": tracker.logs}
